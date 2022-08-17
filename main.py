import sys
import os

from PyQt6.QtCore import Qt, QDateTime
from PyQt6.QtWidgets import QApplication
from PyQt6.QtWidgets import QLabel
from PyQt6.QtWidgets import (
    QWidget,
    QFileDialog,
    QMainWindow,
    QMenuBar,
    QMenu,
    QGridLayout,
    QLineEdit,
    QDateEdit,
    QPushButton,
    QMessageBox,
)

from openpyxl import load_workbook


class Window(QMainWindow):
    def __init__(self, parent=None):
        """Initializer."""
        super().__init__(parent)
        self.setWindowTitle("ParkProReakt Patienten Randomisierung")
        self.resize(400, 200)
        self._createMenuBar()
        self._check_for_config()
        self._createForm()

    def _createMenuBar(self):
        menuBar = QMenuBar()
        self.setMenuBar(menuBar)
        # Creating menus using a QMenu object
        fileMenu = QMenu(" &File", self)
        menuBar.addMenu(fileMenu)
        # Creating menus using a title
        fileMenu.addAction("Exceltabelle auswählen", self._selectFile)

    def _createForm(self):
        # Widget und Layout
        self.form_widget = QWidget(parent=self)
        self.grid_layout = QGridLayout()
        self.form_widget.setLayout(self.grid_layout)

        # Formular Label und Felder
        name_label = QLabel(self.form_widget)
        name_label.setText("Name: ")
        self.name_value = QLineEdit()

        birthday_label = QLabel(self.form_widget)
        birthday_label.setText("Birthday: ")
        self.birthday_value = QDateEdit(calendarPopup=True)
        self.birthday_value.setDateTime(QDateTime.currentDateTime())
        self.birthday_value.setDisplayFormat("dd.MM.yyyy")

        height_label = QLabel(self.form_widget)
        height_label.setText("Height (in cm): ")
        self.height_value = QLineEdit()

        weight_label = QLabel(self.form_widget)
        weight_label.setText("Weight (in kg): ")
        self.weight_value = QLineEdit()

        # Buttons
        self.clearButton = QPushButton()
        self.clearButton.setText("Clear")

        self.addButton = QPushButton()
        self.addButton.setText("Add Patient")

        self.addButton.clicked.connect(self.addPatient)
        self.clearButton.clicked.connect(self.clearForm)

        # Formularlayout
        self.grid_layout.addWidget(name_label, 0, 0)
        self.grid_layout.addWidget(self.name_value, 0, 1)
        self.grid_layout.addWidget(birthday_label, 1, 0)
        self.grid_layout.addWidget(self.birthday_value, 1, 1)
        self.grid_layout.addWidget(height_label, 2, 0)
        self.grid_layout.addWidget(self.height_value, 2, 1)
        self.grid_layout.addWidget(weight_label, 3, 0)
        self.grid_layout.addWidget(self.weight_value, 3, 1)
        self.grid_layout.addWidget(self.clearButton, 4, 0)
        self.grid_layout.addWidget(self.addButton, 4, 1)

        self.setCentralWidget(self.form_widget)
        self.grid_layout.setAlignment(
            Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
        )

    def _selectFile(self):
        self.excel_file = QFileDialog.getOpenFileName(
            self, "Exceltabelle auswählen", "/", filter="Excel files (*.xls *.xlsx)"
        )[0]
        with open(os.getcwd() + os.path.sep + "config.ini", "w") as config_file:
            config_file.write(self.excel_file)

    def _check_for_config(self):
        if "config.ini" in os.listdir(os.getcwd()):
            with open(os.getcwd() + os.path.sep + "config.ini", "r") as config_file:
                self.excel_file = config_file.read()

    def clearForm(self):
        for item in self.form_widget.children():
            if type(item) == QLineEdit:
                item.setText("")
            elif type(item) == QDateEdit:
                item.setDateTime(QDateTime.currentDateTime())
            else:
                continue

    def addPatient(self):
        workbook = load_workbook(self.excel_file)
        sheet = workbook.active

        current_patients = sheet.max_row - 1

        dialog_response = self._showDialog(current_patients + 1)

        if dialog_response == QMessageBox.StandardButton.Ok:
            sheet["A" + str(current_patients + 2)] = self.name_value.text()
            sheet[
                "B" + str(current_patients + 2)
            ] = self.birthday_value.dateTime().toString("dd.MM.yyyy")
            sheet["C" + str(current_patients + 2)] = self.height_value.text()
            sheet["D" + str(current_patients + 2)] = self.weight_value.text()

            workbook.save(self.excel_file)

            self.clearForm()

        else:
            return

    def _showDialog(self, number_of_patients):
        msg = QMessageBox()
        msg.setWindowTitle("Adding Patient")
        msg.setText(
            "Are you sure you want to add this data as the %d. patient too the study?"
            % number_of_patients
        )
        msg.setStandardButtons(
            QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel
        )

        return msg.exec()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = Window()
    win.show()
    sys.exit(app.exec())
